import argparse
import csv
import os
from pathlib import Path

from openpyxl import load_workbook


def convert_xlsx_to_tsv(xlsx_path: str, tsv_path: str | None = None, sheet: str | None = None) -> str:
    src = Path(xlsx_path)
    if not src.exists():
        raise FileNotFoundError(str(src))

    dst = Path(tsv_path) if tsv_path else src.with_suffix(".tsv")

    wb = load_workbook(filename=str(src), read_only=True, data_only=True)
    try:
        if sheet is None:
            ws = wb[wb.sheetnames[0]]
        else:
            if sheet not in wb.sheetnames:
                raise ValueError(f"Sheet '{sheet}' not found. Available: {wb.sheetnames}")
            ws = wb[sheet]

        dst.parent.mkdir(parents=True, exist_ok=True)
        with open(dst, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(
                f,
                delimiter="\t",
                lineterminator="\n",
                quoting=csv.QUOTE_MINIMAL,
            )
            for row in ws.iter_rows(values_only=True):
                # Normalize None to empty string to keep column alignment in TSV
                writer.writerow(["" if v is None else v for v in row])
    finally:
        wb.close()

    return str(dst)


def main() -> None:
    parser = argparse.ArgumentParser(description="Convert .xlsx to tab-separated .tsv (streaming, low-memory).")
    parser.add_argument("--file", required=True, help="Path to .xlsx file")
    parser.add_argument("--out", default=None, help="Optional output .tsv path")
    parser.add_argument("--sheet", default=None, help="Optional sheet name (default: first sheet)")
    args = parser.parse_args()

    out = convert_xlsx_to_tsv(args.file, args.out, args.sheet)
    print(out)


if __name__ == "__main__":
    main()

